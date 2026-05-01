from __future__ import annotations

import math
import re
import warnings
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# All DSP export sheets use header row index 5 (0-based), data from row 6.
# pandas header= parameter is also 0-based, so header=5.
# ---------------------------------------------------------------------------
DEFAULT_HEADER_ROW = 5

# ---------------------------------------------------------------------------
# Sheet prefixes to load via pandas (fast path).
# The header sheet (01) is still read via openpyxl for cell-level access.
# ---------------------------------------------------------------------------
PANDAS_SHEETS = [
    '02_DSP_Date_Range_KPIs',
    '03_DSP_L24M_Monthly_Performance',
    '04_DSP_Monthly_YoY_Comparison',
    '05_DSP_Yearly_KPIs__Current_vs_',
    '06_DSP_Order_Report',
    '07_DSP_Spend_by_Strategy_&_Funn',
    '08_DSP_ASIN_Level_Report',
    '09_DSP_LineItem_Report',
    '10_DSP_vs_PPC_Comparison',
    '11_Gong_Call_Insights',
    '12_Client_Journey_Insights',
    '13_Client_Success_Insights',
    '14_DSP_Project_on_SF',
    '15_Customer_Journey_Funnel_Segm',
    '16_Customer_Journey_Funnel_Segm',
]


@dataclass
class DSPContext:
    path: str
    # ── Account identity ──────────────────────────────────────────────────
    hash_name:    str = ''
    tenant_id:    str = ''
    advertiser_id: str = ''
    window_start: object = None
    window_end:   object = None
    downloaded:   object = None
    window_days:  Optional[int] = None
    ref_date:     object = None

    # ── 02: KPI summary ───────────────────────────────────────────────────
    df02: Optional[pd.DataFrame] = None

    # ── 03: L24M monthly trend ────────────────────────────────────────────
    df03: Optional[pd.DataFrame] = None

    # ── 04: YoY comparison ────────────────────────────────────────────────
    df04: Optional[pd.DataFrame] = None

    # ── 05: Yearly KPIs ───────────────────────────────────────────────────
    df05: Optional[pd.DataFrame] = None

    # ── 06: Order report ──────────────────────────────────────────────────
    df06: Optional[pd.DataFrame] = None

    # ── 07: Spend by strategy & funnel ───────────────────────────────────
    df07: Optional[pd.DataFrame] = None

    # ── 08: ASIN level ────────────────────────────────────────────────────
    df08: Optional[pd.DataFrame] = None

    # ── 09: Line item report ─────────────────────────────────────────────
    df09: Optional[pd.DataFrame] = None

    # ── 10: DSP vs PPC ───────────────────────────────────────────────────
    df10: Optional[pd.DataFrame] = None

    # ── 11: Gong calls ────────────────────────────────────────────────────
    df11: Optional[pd.DataFrame] = None

    # ── 12: Client Journey Map ────────────────────────────────────────────
    df12: Optional[pd.DataFrame] = None

    # ── 13: Client Success Insights ───────────────────────────────────────
    df13: Optional[pd.DataFrame] = None

    # ── 14: DSP Project on SF ─────────────────────────────────────────────
    df14: Optional[pd.DataFrame] = None

    # ── 15: Customer Journey Funnel Segment Mapping ───────────────────────
    df15: Optional[pd.DataFrame] = None

    # ── 16: Customer Journey Funnel Segment Keywords ──────────────────────
    df16: Optional[pd.DataFrame] = None

    # ── Derived / cached scalars ──────────────────────────────────────────
    daily_target:   Optional[float] = None
    target_roas:    Optional[float] = None
    target_acos:    Optional[float] = None   # Target_ACoS__c from sheet 14
    target_tacos:   Optional[float] = None   # Target_TACoS__c from sheet 14
    at_risk:        Optional[bool]  = None
    risk_notes:     str = ''
    cs_notes:       str = ''
    project_status: str = ''

    # ── Constraints & KPI (from sheet 13) ────────────────────────────────
    primary_kpi:       str   = 'ROAS'          # Primary_Spend_KPI__c
    acos_constraint:   Optional[float] = None  # ACOS_Constraint__c (as %)
    tacos_constraint:  Optional[float] = None  # TACOS_Constraint__c (as %)
    budget_constraint: Optional[float] = None  # Monthly_Budget__c

    primary_objective:   str = ''
    current_challenges:  str = ''
    near_term:           str = ''
    churn_risk:          str = ''
    account_risk_score:  object = None
    csi_last_modified:   object = None

    journey_stage:    str = ''
    journey_strategy: str = ''

    last_call_date:   object = None
    days_since_call:  Optional[int] = None
    calls_l90d:       int = 0

    roas:         Optional[float] = None
    ad_spend:     Optional[float] = None
    ad_sales:     Optional[float] = None
    add_to_cart:  Optional[float] = None
    dpv:          Optional[float] = None
    ntb_rate:     Optional[float] = None
    ntb_purchases: Optional[float] = None
    impressions:  Optional[float] = None

    yoy_sales_growth: Optional[float] = None
    yoy_spend_growth: Optional[float] = None

    upper_pct:  Optional[float] = None
    mid_pct:    Optional[float] = None
    lower_pct:  Optional[float] = None
    upper_roas: Optional[float] = None
    lower_roas: Optional[float] = None

    dsp_spend_total: Optional[float] = None
    ppc_spend_total: Optional[float] = None

    ctr_trend:  list = field(default_factory=list)
    cvr_trend:  list = field(default_factory=list)
    cpm_trend:  list = field(default_factory=list)
    dpvr_trend: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_text(v) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ''
    return str(v).replace('&#39;', "'").strip()


def to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        return float(v)
    s = clean_text(v)
    if not s or s.lower() in {'nan', 'none', 'null', '-'}:
        return None
    s = s.replace('$', '').replace(',', '').replace('%', '').strip()
    try:
        return float(s)
    except Exception:
        return None


def money_str(v: Optional[float]) -> str:
    if v is None:
        return 'Not documented'
    return f'${v:,.0f}'


def roas_str(v: Optional[float]) -> str:
    if v is None:
        return 'Not documented'
    return f'{v:.2f}x'


def pct_str(v: Optional[float], decimals: int = 1) -> str:
    if v is None:
        return 'Not documented'
    return f'{v:.{decimals}f}%'


def trim(s: str, n: int = 280) -> str:
    s = re.sub(r'\s+', ' ', s or '').strip()
    return s if len(s) <= n else s[: n - 1].rstrip() + '…'


def _find_col(df: pd.DataFrame, candidates: list) -> Optional[str]:
    norm = {str(c).strip().lower().replace(' ', '').replace('_', ''): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower().replace(' ', '').replace('_', '')
        if key in norm:
            return norm[key]
    return None


# ---------------------------------------------------------------------------
# Header parser — reads 01_DSP_Advertiser sheet via openpyxl (cell access needed)
# ---------------------------------------------------------------------------

def _parse_header_from_path(path: str) -> dict:
    """Open workbook once just to read the 4-row header from 01_DSP_Advertiser."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_name = next((n for n in wb.sheetnames if n.startswith('01_DSP_Advertiser')), None)
        if sheet_name is None:
            raise ValueError('Sheet 01_DSP_Advertiser_Name not found in DSP export.')
        ws = wb[sheet_name]
        a1 = clean_text(ws['A1'].value)
        a2 = clean_text(ws['A2'].value)
        a3 = clean_text(ws['A3'].value)
        a4 = clean_text(ws['A4'].value)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    hash_name = re.sub(r'\s*-\s*DSP Advertiser.*$', '', a1, flags=re.I).strip()

    m2 = re.search(r'Tenant ID:\s*(.*?)\s*\|\s*Advertiser ID:\s*(.*)$', a2)
    tenant = m2.group(1).strip() if m2 else ''
    adv_id = m2.group(2).strip() if m2 else ''

    m3 = re.search(r'Date Range:\s*([0-9\-]+)\s*to\s*([0-9\-]+)', a3)
    start = datetime.strptime(m3.group(1), '%Y-%m-%d').date() if m3 else None
    end   = datetime.strptime(m3.group(2), '%Y-%m-%d').date() if m3 else None

    dl_raw = re.sub(r'^Downloaded:\s*', '', a4).strip()
    try:
        downloaded = datetime.strptime(dl_raw, '%Y-%m-%d %H:%M:%S')
    except Exception:
        downloaded = None

    return {
        'hash_name':     hash_name,
        'tenant_id':     tenant,
        'advertiser_id': adv_id,
        'window_start':  start,
        'window_end':    end,
        'downloaded':    downloaded,
        'window_days':   (end - start).days + 1 if start and end else None,
        'ref_date':      downloaded.date() if downloaded else end,
    }


# ---------------------------------------------------------------------------
# Fast sheet loader — pd.ExcelFile opened once, all sheets read via pandas
# ---------------------------------------------------------------------------

def _load_sheets_pandas(path: str) -> dict:
    """
    Open the file once with pd.ExcelFile and read all PANDAS_SHEETS in one pass.
    pandas read_excel is 5-10x faster than openpyxl iter_rows for the same data.
    Returns dict of prefix -> DataFrame (or None if sheet not found).
    """
    xls = pd.ExcelFile(path)
    available = xls.sheet_names

    result = {}
    for prefix in PANDAS_SHEETS:
        match = next((n for n in available if n.startswith(prefix)), None)
        if match is None:
            result[prefix] = None
            continue
        try:
            df = pd.read_excel(xls, sheet_name=match, header=DEFAULT_HEADER_ROW)
            # Drop fully-empty rows
            df = df.dropna(how='all')
            result[prefix] = df if not df.empty else None
        except Exception:
            result[prefix] = None

    return result


# ---------------------------------------------------------------------------
# Gong helpers
# ---------------------------------------------------------------------------

def _gong_stats(df: Optional[pd.DataFrame], ref_date) -> tuple:
    if df is None or df.empty:
        return None, None, 0
    col = _find_col(df, ['Gong__Call_Start__c', 'Gong__Call_End__c'])
    if col is None:
        return None, None, 0
    dates = pd.to_datetime(df[col], errors='coerce').dropna().sort_values(ascending=False)
    if dates.empty:
        return None, None, 0
    last = dates.iloc[0]
    ref = pd.Timestamp(ref_date) if ref_date else pd.Timestamp.now()
    days = int((ref - last).days)
    cutoff = ref - pd.Timedelta(days=90)
    count = int((dates >= cutoff).sum())
    return last.date(), days, count


# ---------------------------------------------------------------------------
# Funnel aggregation
# ---------------------------------------------------------------------------

UPPER_LABELS = {'upper funnel', 'awareness', 'prospecting'}
LOWER_LABELS = {'lower funnel', 'retargeting', 'rmkt', 'performance'}
MID_LABELS   = {'mid funnel', 'consideration', 'middle funnel'}


def _funnel_label(s: str) -> str:
    s = (s or '').strip().lower()
    if any(k in s for k in LOWER_LABELS):
        return 'lower'
    if any(k in s for k in UPPER_LABELS):
        return 'upper'
    if any(k in s for k in MID_LABELS):
        return 'mid'
    return 'other'


def _compute_funnel_splits(df07: Optional[pd.DataFrame]) -> tuple:
    if df07 is None or df07.empty:
        return None, None, None, None, None
    spend_col  = _find_col(df07, ['AdSpend', 'adspend'])
    sales_col  = _find_col(df07, ['AdSales', 'adsales'])
    funnel_col = _find_col(df07, ['FunnelStage', 'funnelstage'])
    if spend_col is None or funnel_col is None:
        return None, None, None, None, None

    df = df07.copy()
    df['_spend'] = pd.to_numeric(df[spend_col], errors='coerce').fillna(0)
    df['_sales'] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0) if sales_col else 0
    df['_layer'] = df[funnel_col].apply(lambda x: _funnel_label(clean_text(x)))

    total_spend = df['_spend'].sum()
    if total_spend == 0:
        return 0.0, 0.0, 0.0, None, None

    by_layer = df.groupby('_layer')[['_spend', '_sales']].sum()

    def _pct(layer):
        return float(by_layer.loc[layer, '_spend'] / total_spend * 100) if layer in by_layer.index else 0.0

    def _roas(layer):
        if layer not in by_layer.index:
            return None
        s = float(by_layer.loc[layer, '_spend'])
        r = float(by_layer.loc[layer, '_sales'])
        return r / s if s > 0 else None

    return _pct('upper'), _pct('mid'), _pct('lower'), _roas('upper'), _roas('lower')


# ---------------------------------------------------------------------------
# Internal helpers for row-level scalar extraction
# ---------------------------------------------------------------------------

def _find_col_val(row, candidates: list):
    if row is None:
        return None
    for c in candidates:
        key = c.strip().lower().replace('_', '').replace(' ', '')
        for idx in row.index:
            if str(idx).strip().lower().replace('_', '').replace(' ', '') == key:
                v = row[idx]
                return None if (isinstance(v, float) and math.isnan(v)) else v
    return None


def _row_val(row, candidates: list):
    return _find_col_val(row, candidates)


def _bool_val(v) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {'true', '1', 'yes'}:
        return True
    if s in {'false', '0', 'no'}:
        return False
    return None


def trend_direction(values: list) -> str:
    vals = [v for v in values if v is not None]
    if len(vals) < 2:
        return 'insufficient data'
    diffs = [vals[i+1] - vals[i] for i in range(len(vals)-1)]
    neg = sum(1 for d in diffs if d < -0.0001)
    pos = sum(1 for d in diffs if d > 0.0001)
    if neg == len(diffs):
        return 'declining'
    if pos == len(diffs):
        return 'improving'
    return 'stable'


def _latest_row_by_modstamp(df: pd.DataFrame) -> pd.Series:
    """Return the row with the most recent SystemModstamp. Falls back to iloc[0]."""
    mod_col = next(
        (c for c in df.columns if 'systemmod' in str(c).lower() or 'modstamp' in str(c).lower()),
        None
    )
    if mod_col:
        try:
            df2 = df.copy()
            df2['_ts'] = pd.to_datetime(df2[mod_col], errors='coerce')
            valid = df2.dropna(subset=['_ts'])
            if not valid.empty:
                return valid.loc[valid['_ts'].idxmax()].drop(labels=['_ts'])
        except Exception:
            pass
    return df.iloc[0]


# ---------------------------------------------------------------------------
# Main loader
# ---------------------------------------------------------------------------

def load_dsp_context(path: str) -> DSPContext:
    # ── Step 1: Read header (openpyxl, read_only — only the 01 sheet) ─────
    h = _parse_header_from_path(path)
    ctx = DSPContext(path=path, **h)

    # ── Step 2: Load all data sheets via pandas (fast) ────────────────────
    sheets = _load_sheets_pandas(path)

    ctx.df02 = sheets.get('02_DSP_Date_Range_KPIs')
    ctx.df03 = sheets.get('03_DSP_L24M_Monthly_Performance')
    ctx.df04 = sheets.get('04_DSP_Monthly_YoY_Comparison')
    ctx.df05 = sheets.get('05_DSP_Yearly_KPIs__Current_vs_')
    ctx.df06 = sheets.get('06_DSP_Order_Report')
    ctx.df07 = sheets.get('07_DSP_Spend_by_Strategy_&_Funn')
    ctx.df08 = sheets.get('08_DSP_ASIN_Level_Report')
    ctx.df09 = sheets.get('09_DSP_LineItem_Report')
    ctx.df10 = sheets.get('10_DSP_vs_PPC_Comparison')
    ctx.df11 = sheets.get('11_Gong_Call_Insights')
    ctx.df12 = sheets.get('12_Client_Journey_Insights')
    ctx.df13 = sheets.get('13_Client_Success_Insights')
    ctx.df14 = sheets.get('14_DSP_Project_on_SF')
    ctx.df15 = sheets.get('15_Customer_Journey_Funnel_Segm')
    ctx.df16 = sheets.get('16_Customer_Journey_Funnel_Segm')

    # ── Step 3: Derive scalars (no I/O, pure computation) ─────────────────

    # 14 — DSP Project
    if ctx.df14 is not None and not ctx.df14.empty:
        adv_col = _find_col(ctx.df14, ['Advertiser_ID__c', 'advertiserid'])
        df14 = ctx.df14
        if adv_col and ctx.advertiser_id:
            mask = df14[adv_col].astype(str).str.strip() == str(ctx.advertiser_id).strip()
            match = df14[mask]
            proj_row = match.iloc[0] if not match.empty else df14.iloc[0]
        else:
            proj_row = df14.iloc[0]

        ctx.daily_target   = to_float(_find_col_val(proj_row, ['daily_target_spend__c']))
        ctx.target_roas    = to_float(_find_col_val(proj_row, ['Target_ROAS__c']))
        ctx.target_acos    = to_float(_find_col_val(proj_row, ['Target_ACoS__c']))
        ctx.target_tacos   = to_float(_find_col_val(proj_row, ['Target_TACoS__c']))
        ctx.at_risk        = _bool_val(_find_col_val(proj_row, ['At_Risk__c']))
        ctx.risk_notes     = clean_text(_find_col_val(proj_row, ['Risk_Reason_Notes__c']))
        ctx.cs_notes       = clean_text(_find_col_val(proj_row, ['CS_Notes__c']))
        ctx.project_status = clean_text(_find_col_val(proj_row, ['Project_Status__c']))

    # 13 — Client Success Insights (latest row by SystemModstamp)
    if ctx.df13 is not None and not ctx.df13.empty:
        r = _latest_row_by_modstamp(ctx.df13)
        ctx.primary_objective  = clean_text(_row_val(r, ['Primary_Objective__c']))
        ctx.current_challenges = clean_text(_row_val(r, ['Current_Challenges__c']))
        ctx.near_term          = clean_text(_row_val(r, ['Near_Term_3_Month_Considerations__c']))
        ctx.churn_risk         = clean_text(_row_val(r, ['CSM_Churn_Risk__c']))
        ctx.account_risk_score = _row_val(r, ['Account_Risk_Score__c'])
        ctx.csi_last_modified  = _row_val(r, ['LastModifiedDate'])

        # Constraints and primary KPI — mirrors Amazon Health reader
        raw_kpi = clean_text(_row_val(r, ['Primary_Spend_KPI__c'])).upper()
        ctx.primary_kpi = raw_kpi if raw_kpi in ('ACOS', 'TACOS', 'ROAS') else 'ROAS'

        acos_raw = to_float(_row_val(r, ['ACOS_Constraint__c']))
        tacos_raw = to_float(_row_val(r, ['TACOS_Constraint__c']))
        # Store as plain % value (e.g. 12 means 12%) — writer formats it
        ctx.acos_constraint  = acos_raw
        ctx.tacos_constraint = tacos_raw
        ctx.budget_constraint = to_float(_row_val(r, ['Monthly_Budget__c']))

    # 12 — Client Journey Map
    if ctx.df12 is not None and not ctx.df12.empty:
        r = ctx.df12.iloc[0]
        for stage_n, strat_n in [('StatusS1__c', 'StrategyS1__c'), ('StatusS2__c', 'StrategyS2__c'),
                                  ('StatusS3__c', 'StrategyS3__c'), ('StatusS4__c', 'StrategyS4__c')]:
            sv = clean_text(_row_val(r, [stage_n]))
            if sv and sv.lower() not in {'', 'none', 'nan', 'not started'}:
                ctx.journey_stage    = sv
                ctx.journey_strategy = clean_text(_row_val(r, [strat_n]))
                break

    # 11 — Gong
    ctx.last_call_date, ctx.days_since_call, ctx.calls_l90d = _gong_stats(ctx.df11, ctx.ref_date)

    # 02 — KPI scalars
    if ctx.df02 is not None and not ctx.df02.empty:
        r = ctx.df02.iloc[0]
        ctx.roas          = to_float(_row_val(r, ['ROAS']))
        ctx.ad_spend      = to_float(_row_val(r, ['AdSpend']))
        ctx.ad_sales      = to_float(_row_val(r, ['AdSales']))
        ctx.add_to_cart   = to_float(_row_val(r, ['AddToCart']))
        ctx.dpv           = to_float(_row_val(r, ['DPV']))
        ctx.impressions   = to_float(_row_val(r, ['Impressions']))
        ntb  = to_float(_row_val(r, ['NTBPurchases']))
        conv = to_float(_row_val(r, ['Conversions']))
        ctx.ntb_purchases = ntb
        if ntb is not None and conv is not None and conv > 0:
            ctx.ntb_rate = ntb / conv * 100

    # 04 — YoY growth
    if ctx.df04 is not None and not ctx.df04.empty:
        col_yoy_sales = _find_col(ctx.df04, ['YoY_AdSales'])
        col_yoy_spend = _find_col(ctx.df04, ['YoY_AdSpend'])
        if col_yoy_sales:
            vals = pd.to_numeric(ctx.df04[col_yoy_sales], errors='coerce').dropna()
            ctx.yoy_sales_growth = float(vals.iloc[-1]) if not vals.empty else None
        if col_yoy_spend:
            vals = pd.to_numeric(ctx.df04[col_yoy_spend], errors='coerce').dropna()
            ctx.yoy_spend_growth = float(vals.iloc[-1]) if not vals.empty else None

    # 07 — Funnel splits
    ctx.upper_pct, ctx.mid_pct, ctx.lower_pct, ctx.upper_roas, ctx.lower_roas = _compute_funnel_splits(ctx.df07)

    # 10 — DSP vs PPC
    if ctx.df10 is not None and not ctx.df10.empty:
        channel_col = _find_col(ctx.df10, ['Channel', 'channel'])
        spend_col   = _find_col(ctx.df10, ['AdSpendUSD', 'adspend'])
        if channel_col and spend_col:
            df10 = ctx.df10.copy()
            df10['_spend'] = pd.to_numeric(df10[spend_col], errors='coerce').fillna(0)
            df10['_ch'] = df10[channel_col].astype(str).str.upper().str.strip()
            ctx.dsp_spend_total = float(df10[df10['_ch'] == 'DSP']['_spend'].sum())
            ctx.ppc_spend_total = float(df10[df10['_ch'].isin({'PPC', 'SP', 'SB', 'SD', 'SPONSORED'})]['_spend'].sum())
            if ctx.ppc_spend_total == 0:
                ctx.ppc_spend_total = float(df10[df10['_ch'] != 'DSP']['_spend'].sum())

    # 03 — Trend lists (last 3 months)
    if ctx.df03 is not None and not ctx.df03.empty:
        df3 = ctx.df03.copy()
        month_col = _find_col(df3, ['Month', 'month'])
        if month_col:
            df3['_month'] = pd.to_datetime(df3[month_col], errors='coerce')
            df3 = df3.dropna(subset=['_month']).sort_values('_month')
        last3 = df3.tail(3)

        def _trend(col_candidates):
            col = _find_col(last3, col_candidates)
            if col is None:
                return []
            return [to_float(v) for v in last3[col].tolist()]

        ctx.ctr_trend = _trend(['CTR', 'ctr'])
        ctx.cvr_trend = _trend(['CR', 'CVR', 'ConversionRate', 'Conversions'])
        ctx.cpm_trend = _trend(['CPM', 'cpm'])

        dpv_col  = _find_col(last3, ['DPV'])
        impr_col = _find_col(last3, ['Impressions'])
        if dpv_col and impr_col:
            dpvr_list = []
            for _, row in last3.iterrows():
                d = to_float(row[dpv_col])
                i = to_float(row[impr_col])
                dpvr_list.append(d / i * 100 if d is not None and i and i > 0 else None)
            ctx.dpvr_trend = dpvr_list

    return ctx
