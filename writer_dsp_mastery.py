from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from reader_databricks_dsp import clean_text, money_str, roas_str


def write_dsp_mastery_output(template_path, output_path, results, ctx):
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Account Mastery_Analysis']
    ws_ref  = wb['Account Mastery_Reference']

    # ── Analysis tab header ───────────────────────────────────────────────
    ws_main['A1'] = f"{ctx.hash_name} — Account Mastery Analysis"
    ws_main['B3'] = f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.advertiser_id}"
    if ctx.window_start and ctx.window_end and ctx.window_days:
        ws_main['B4'] = f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)"
    if ctx.downloaded:
        ws_main['B5'] = ctx.downloaded
        ws_main['B5'].number_format = 'yyyy-mm-dd hh:mm:ss'

    # ── Reference tab: write STATUS, What We Saw, Why It Matters ─────────
    cid_to_row = {}
    for row in range(2, ws_ref.max_row + 1):
        cid = clean_text(ws_ref[f'B{row}'].value).upper()
        if cid:
            cid_to_row[cid] = row

    for cid, res in results.items():
        if cid not in cid_to_row:
            print(f'[writer_dsp_mastery] WARNING: {cid} not found in reference sheet — skipping.')
            continue
        rr = cid_to_row[cid]
        ws_ref[f'D{rr}'] = res.status
        ws_ref[f'H{rr}'] = res.what
        ws_ref[f'I{rr}'] = res.why
        for cell_ref in [f'H{rr}', f'I{rr}']:
            ws_ref[cell_ref].alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass
