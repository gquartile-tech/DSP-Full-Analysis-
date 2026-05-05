from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

from reader_databricks_dsp import clean_text


def _safe_write(ws, cell_ref: str, value, number_format: str = None,
                alignment: Alignment = None):
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        from openpyxl.utils import range_boundaries
        to_remove = None
        for rng in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))
            cr = ws[cell_ref]
            if min_row <= cr.row <= max_row and min_col <= cr.column <= max_col:
                to_remove = rng
                break
        if to_remove:
            ws.unmerge_cells(str(to_remove))
        cell = ws[cell_ref]

    cell.value = value
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment


def write_dsp_strategy_output(template_path, output_path, results, ctx):
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Account Strategy_Analysis']
    ws_ref  = wb['Account Strategy_Reference']

    # ── Analysis tab header ───────────────────────────────────────────────
    _safe_write(ws_main, 'A1', f"{ctx.hash_name} — Account Strategy Analysis")
    _safe_write(ws_main, 'B3', f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.advertiser_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, 'B4', f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        _safe_write(ws_main, 'B5', ctx.downloaded, number_format='yyyy-mm-dd hh:mm:ss')

    # ── Reference tab: STATUS, What We Saw, Why It Matters ───────────────
    cid_to_row = {}
    for row in range(2, ws_ref.max_row + 1):
        cell = ws_ref[f'B{row}']
        raw = '' if isinstance(cell, MergedCell) else clean_text(cell.value).upper()
        if raw:
            cid_to_row[raw] = row

    wrap_top = Alignment(wrap_text=True, vertical='top')

    for cid, res in results.items():
        if cid not in cid_to_row:
            print(f'[writer_dsp_strategy] WARNING: {cid} not found in reference sheet — skipping.')
            continue
        rr = cid_to_row[cid]
        _safe_write(ws_ref, f'D{rr}', res.status)
        _safe_write(ws_ref, f'H{rr}', res.what,  alignment=wrap_top)
        _safe_write(ws_ref, f'I{rr}', res.why,   alignment=wrap_top)

    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass
