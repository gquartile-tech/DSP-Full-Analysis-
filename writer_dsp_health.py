from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from reader_databricks_dsp import clean_text, to_float


def _fmt_pct(v, decimals: int = 1) -> str:
    """Format a raw % value (e.g. 12.0 -> '12.0%'). Returns 'Not set' if None."""
    if v is None:
        return 'Not set'
    try:
        return f'{float(v):.{decimals}f}%'
    except Exception:
        return 'Not set'


def _fmt_money(v) -> str:
    if v is None:
        return 'Not set'
    try:
        return f'${float(v):,.0f}'
    except Exception:
        return 'Not set'


def write_dsp_health_output(template_path, output_path, results, ctx):
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb['Account Health_Analysis']
    ws_ref  = wb['Account Health_Reference']

    # ── Analysis tab header ───────────────────────────────────────────────
    ws_main['A1'] = f"{ctx.hash_name} — Account Health Analysis"
    ws_main['B3'] = f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.advertiser_id}"
    if ctx.window_start and ctx.window_end and ctx.window_days:
        ws_main['B4'] = f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)"
    if ctx.downloaded:
        ws_main['B5'] = ctx.downloaded
        ws_main['B5'].number_format = 'yyyy-mm-dd hh:mm:ss'

    # ── Constraints & Primary KPI block (mirrors Amazon Health writer) ────
    ws_main['B9']  = _fmt_pct(ctx.acos_constraint)
    ws_main['B10'] = _fmt_pct(ctx.tacos_constraint)
    ws_main['B11'] = _fmt_money(ctx.budget_constraint)
    ws_main['E10'] = getattr(ctx, 'primary_kpi', 'ROAS')

    # ── Reference tab: STATUS, What We Saw, Why It Matters ───────────────
    cid_to_row = {}
    for row in range(2, ws_ref.max_row + 1):
        cid = clean_text(ws_ref[f'B{row}'].value).upper()
        if cid:
            cid_to_row[cid] = row

    for cid, res in results.items():
        if cid not in cid_to_row:
            print(f'[writer_dsp_health] WARNING: {cid} not found in reference sheet — skipping.')
            continue
        rr = cid_to_row[cid]
        ws_ref[f'D{rr}'] = res.status
        ws_ref[f'H{rr}'] = res.what
        ws_ref[f'I{rr}'] = res.why
        ws_ref[f'J{rr}'] = res.action
        for cell_ref in [f'H{rr}', f'I{rr}', f'J{rr}']:
            ws_ref[cell_ref].alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass
